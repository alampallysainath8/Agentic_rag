"""
Batch Evaluator
===============
Loads a test dataset (Excel or JSON), runs DeepEval RAG metrics against
every test case, and writes results back to a new Excel workbook.

Column contract for the input file
------------------------------------
Required: ``input``, ``expected_output``, ``retrieval_context``
Optional: ``id``, ``pattern``, ``category``, ``notes``

``retrieval_context`` can be:
  - A JSON array string:  ``["chunk1", "chunk2"]``
  - Newline-separated text chunks (one per line)
  - A single flat string (treated as one chunk)

Usage
-----
    from src.evaluation.batch_evaluator import BatchEvaluator

    rpt = BatchEvaluator().run_from_excel("path/to/hr_qa_dataset.xlsx")
    print(rpt.summary())
"""
from __future__ import annotations

import json
import logging
import re
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pandas as pd

from .evaluator import EvaluationResult, RAGEvaluatorService

logger = logging.getLogger(__name__)


# ── helpers ────────────────────────────────────────────────────────────────────

def _parse_context(value: Any) -> list[str]:
    """Normalise a retrieval_context cell to a list of strings."""
    if isinstance(value, list):
        return [str(v) for v in value if v]
    s = str(value).strip()
    if not s or s.lower() in ("nan", "none", ""):
        return []
    # try JSON array
    if s.startswith("["):
        try:
            parsed = json.loads(s)
            if isinstance(parsed, list):
                return [str(p) for p in parsed if p]
        except (json.JSONDecodeError, ValueError):
            pass
    # newline-delimited chunks
    lines = [ln.strip() for ln in s.splitlines() if ln.strip()]
    return lines if lines else [s]


# ── result container ───────────────────────────────────────────────────────────

@dataclass
class BatchReport:
    total:   int
    passed:  int
    failed:  int
    errored: int
    results: list[dict[str, Any]] = field(default_factory=list)
    output_path: str = ""

    def summary(self) -> str:
        rate = (self.passed / self.total * 100) if self.total else 0
        return (
            f"Batch evaluation complete — {self.total} cases | "
            f"{self.passed} passed ({rate:.1f}%) | "
            f"{self.failed} failed | {self.errored} errored"
        )

    def metric_averages(self) -> dict[str, float]:
        """Average score across all cases for each metric."""
        sums: dict[str, float] = {}
        cnts: dict[str, int]   = {}
        for r in self.results:
            for mk, score in r.get("scores", {}).items():
                sums[mk] = sums.get(mk, 0.0) + score
                cnts[mk] = cnts.get(mk, 0)  + 1
        return {mk: round(sums[mk] / cnts[mk], 4) for mk in sums}


# ── batch evaluator ────────────────────────────────────────────────────────────

class BatchEvaluator:
    """
    Run DeepEval RAG metrics over a test dataset stored in Excel or JSON.

    Parameters
    ----------
    model     : Judge LLM (defaults to ``EVAL_MODEL`` env var or ``gpt-4o-mini``).
    threshold : Pass/fail threshold in [0, 1].
    delay_sec : Seconds to wait between API calls (rate-limit buffer).
    """

    def __init__(
        self,
        model: str | None = None,
        threshold: float = 0.5,
        delay_sec: float = 1.0,
    ) -> None:
        self.service   = RAGEvaluatorService(model=model, threshold=threshold)
        self.delay_sec = delay_sec

    # ── loaders ────────────────────────────────────────────────────────────

    def load_excel(self, path: str | Path) -> pd.DataFrame:
        df = pd.read_excel(str(path), sheet_name=0)
        df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]
        return df

    def load_json(self, path: str | Path) -> pd.DataFrame:
        data = json.loads(Path(path).read_text(encoding="utf-8"))
        df = pd.DataFrame(data)
        df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]
        return df

    # ── core runner ────────────────────────────────────────────────────────

    def run_dataframe(self, df: pd.DataFrame) -> BatchReport:
        """
        Evaluate every row in *df*.  Required columns:
        ``input``, ``expected_output``, ``retrieval_context``.
        """
        required = {"input", "expected_output", "retrieval_context"}
        missing  = required - set(df.columns)
        if missing:
            raise ValueError(f"Dataset is missing columns: {missing}")

        rows: list[dict[str, Any]] = []
        passed = failed = errored = 0

        for i, row in enumerate(df.itertuples(index=False), 1):
            cid      = getattr(row, "id",       f"row-{i}")
            pattern  = getattr(row, "pattern",  "")
            category = getattr(row, "category", "")
            query    = str(getattr(row, "input"))
            expected = str(getattr(row, "expected_output", "") or "")
            ctx_raw  = getattr(row, "retrieval_context", "")
            chunks   = _parse_context(ctx_raw)

            logger.info("Evaluating [%d/%d] %s", i, len(df), cid)

            record: dict[str, Any] = {
                "id":       cid,
                "pattern":  pattern,
                "category": category,
                "input":    query,
                "expected_output": expected,
                "context_chunks":  chunks,
                "scores":   {},
                "reasons":  {},
                "passed":   {},
                "overall_pass": False,
                "error":    None,
            }

            if not chunks:
                record["error"] = "retrieval_context is empty — skipped"
                errored += 1
                rows.append(record)
                continue

            try:
                result: EvaluationResult = self.service.evaluate(
                    query=query,
                    answer=expected,           # use expected as the "answer" to score
                    context_chunks=chunks,
                    expected_output=expected if expected.strip() else None,
                )
                record["scores"]       = result.scores
                record["reasons"]      = result.reasons
                record["passed"]       = result.passed
                record["overall_pass"] = result.overall_pass
                record["eval_id"]      = result.eval_id
                record["timestamp"]    = result.timestamp
                if result.overall_pass:
                    passed += 1
                else:
                    failed += 1
            except Exception as exc:
                logger.warning("Evaluation error for %s: %s", cid, exc)
                record["error"] = str(exc)
                errored += 1

            rows.append(record)

            if i < len(df) and self.delay_sec > 0:
                time.sleep(self.delay_sec)

        return BatchReport(
            total=len(df),
            passed=passed,
            failed=failed,
            errored=errored,
            results=rows,
        )

    # ── convenience entry-points ───────────────────────────────────────────

    def run_from_excel(
        self,
        path: str | Path,
        output_path: str | Path | None = None,
    ) -> BatchReport:
        """Load Excel, run eval, write results Excel, return :class:`BatchReport`."""
        df  = self.load_excel(path)
        rpt = self.run_dataframe(df)
        out = output_path or Path(path).with_name(
            Path(path).stem + "_results.xlsx"
        )
        self.write_results_excel(rpt, out)
        rpt.output_path = str(out)
        return rpt

    def run_from_json(
        self,
        path: str | Path,
        output_path: str | Path | None = None,
    ) -> BatchReport:
        """Load JSON, run eval, write results Excel, return :class:`BatchReport`."""
        df  = self.load_json(path)
        rpt = self.run_dataframe(df)
        out = output_path or Path(path).with_name(
            Path(path).stem + "_results.xlsx"
        )
        self.write_results_excel(rpt, out)
        rpt.output_path = str(out)
        return rpt

    # ── Excel writer ───────────────────────────────────────────────────────

    def write_results_excel(
        self, rpt: BatchReport, path: str | Path
    ) -> None:
        """Write detailed per-case results + summary tab to an Excel workbook."""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, Reference

        METRIC_KEYS = [
            "faithfulness",
            "answer_relevancy",
            "context_relevancy",
            "context_recall",
        ]
        METRIC_LABELS = {
            "faithfulness":      "Faithfulness",
            "answer_relevancy":  "Answer Relevancy",
            "context_relevancy": "Context Relevancy",
            "context_recall":    "Context Recall",
        }

        wb = Workbook()

        # ── helpers ──────────────────────────────────────────────────────
        border_side = Side(border_style="thin", color="BFBFBF")
        cell_border = Border(
            left=border_side, right=border_side,
            top=border_side,  bottom=border_side,
        )
        wrap_top = Alignment(wrap_text=True, vertical="top")

        def _hdr(cell, txt: str, bg: str = "1F3864") -> None:
            cell.value = txt
            cell.font  = Font(bold=True, color="FFFFFF", size=10)
            cell.fill  = PatternFill("solid", fgColor=bg)
            cell.border = cell_border
            cell.alignment = Alignment(horizontal="center",
                                       vertical="center", wrap_text=True)

        def _pass_fill(ok: bool | None) -> PatternFill:
            if ok is True:
                return PatternFill("solid", fgColor="C6EFCE")
            if ok is False:
                return PatternFill("solid", fgColor="FFC7CE")
            return PatternFill("solid", fgColor="FFFFFF")

        # ══ Sheet 1: Results ═══════════════════════════════════════════
        ws = wb.active
        ws.title = "Results"

        COLS = [
            "ID", "Pattern", "Category", "Input",
            "Faithfulness", "Answer Relevancy", "Context Relevancy", "Context Recall",
            "Overall", "Error / Notes",
        ]
        for ci, col in enumerate(COLS, 1):
            _hdr(ws.cell(row=1, column=ci), col)

        for ri, rec in enumerate(rpt.results, 2):
            scores = rec.get("scores", {})
            passed = rec.get("passed", {})
            ok     = rec.get("overall_pass")

            def _cell(ci: int, val: Any, fill: PatternFill | None = None) -> None:
                c = ws.cell(row=ri, column=ci, value=val)
                c.alignment = wrap_top
                c.border    = cell_border
                if fill:
                    c.fill = fill

            _cell(1,  rec.get("id",       ""))
            _cell(2,  rec.get("pattern",  ""))
            _cell(3,  rec.get("category", ""))
            _cell(4,  rec.get("input",    ""))

            for col_i, mk in enumerate(METRIC_KEYS, 5):
                score = scores.get(mk)
                if score is not None:
                    _cell(col_i, round(score, 3),
                          fill=_pass_fill(passed.get(mk)))
                else:
                    _cell(col_i, "—")

            _cell(9,  "✅ PASS" if ok else ("❌ FAIL" if not rec.get("error") else "⚠ ERROR"),
                  fill=_pass_fill(ok if not rec.get("error") else None))
            _cell(10, rec.get("error") or rec.get("reasons", {}).get("faithfulness", ""))

        # column widths
        widths = [10, 12, 12, 50, 16, 18, 18, 16, 12, 40]
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        for ri in range(2, len(rpt.results) + 2):
            ws.row_dimensions[ri].height = 60
        ws.freeze_panes = "A2"

        # ══ Sheet 2: Summary ═══════════════════════════════════════════
        ss = wb.create_sheet("Summary")
        avgs = rpt.metric_averages()

        _hdr(ss["A1"], "Metric",          "2E4057")
        _hdr(ss["B1"], "Avg Score",       "2E4057")
        _hdr(ss["C1"], "Pass Threshold",  "2E4057")
        _hdr(ss["D1"], "Avg Passes",      "2E4057")

        thr = self.service.threshold
        for ri, mk in enumerate(METRIC_KEYS, 2):
            avg  = avgs.get(mk)
            p_ct = sum(
                1 for r in rpt.results
                if r.get("passed", {}).get(mk) is True
            )
            ss.cell(ri, 1, METRIC_LABELS.get(mk, mk)).border = cell_border
            ss.cell(ri, 2, round(avg, 4) if avg is not None else "—").border = cell_border
            ss.cell(ri, 3, thr).border = cell_border
            ss.cell(ri, 4, p_ct).border = cell_border
            if avg is not None:
                ss.cell(ri, 2).fill = _pass_fill(avg >= thr)

        # overall stats
        ss["A7"] = "Total Cases"
        ss["B7"] = rpt.total
        ss["A8"] = "Overall Pass"
        ss["B8"] = rpt.passed
        ss["A9"] = "Overall Fail"
        ss["B9"] = rpt.failed
        ss["A10"]= "Errors"
        ss["B10"]= rpt.errored
        ss["A11"]= "Pass Rate %"
        ss["B11"]= round(rpt.passed / rpt.total * 100, 1) if rpt.total else 0

        for ri in range(7, 12):
            ss.cell(ri, 1).font   = Font(bold=True)
            ss.cell(ri, 1).border = cell_border
            ss.cell(ri, 2).border = cell_border

        ss.column_dimensions["A"].width = 22
        ss.column_dimensions["B"].width = 14
        ss.column_dimensions["C"].width = 16
        ss.column_dimensions["D"].width = 14

        # ══ Sheet 3: Reasons ═══════════════════════════════════════════
        rs = wb.create_sheet("Reasons")
        for ci, lbl in enumerate(["ID", "Input", "Metric", "Score", "Reason"], 1):
            _hdr(rs.cell(row=1, column=ci), lbl)

        ri = 2
        for rec in rpt.results:
            reasons = rec.get("reasons", {})
            scores  = rec.get("scores",  {})
            for mk, reason in reasons.items():
                rs.cell(ri, 1, rec.get("id", "")).border = cell_border
                rs.cell(ri, 2, rec.get("input", "")[:100]).border = cell_border
                rs.cell(ri, 3, METRIC_LABELS.get(mk, mk)).border = cell_border
                rs.cell(ri, 4, round(scores.get(mk, 0), 3)).border = cell_border
                rs.cell(ri, 5, str(reason or "")).border = cell_border
                for ci in range(1, 6):
                    rs.cell(ri, ci).alignment = wrap_top
                ri += 1

        rs.column_dimensions["A"].width = 10
        rs.column_dimensions["B"].width = 40
        rs.column_dimensions["C"].width = 20
        rs.column_dimensions["D"].width = 10
        rs.column_dimensions["E"].width = 70
        rs.freeze_panes = "A2"

        Path(path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(path))
        logger.info("Batch results saved → %s", path)
